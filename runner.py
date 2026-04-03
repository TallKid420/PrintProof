import json
import os
import shutil
from datetime import datetime
from pathlib import Path

import Autofocus
import groq
import takepicture
from Autofocus import show_camera
from Focuser import Focuser
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
JSON_DIR = Path(os.environ.get("IRIS_JSON_DIR", str(BASE_DIR / "Json"))).resolve()
DEFAULT_DOCUMENT_TEMPLATE = [
    'IEEE',
    'In recognition of professional standing',
    'the Officers and Board of Directors of',
    'the IEEE certify that',
    '<expected name>',
    'has been elected to the grade of',
    '<expected title>',
    '<expected date>',
]
FIELD_MAPS = {
    'product': {
        'order_id': ('order id', 'orderid', 'id', 'order number', 'order #'),
        'name': ('name', 'customer name', 'ship name', 'full name'),
        'title': ('title', 'product title', 'item title', 'item', 'product'),
        'date': ('date', 'order date', 'purchase date', 'created at'),
    },
    'shipping': {
        'order_id': ('order id', 'orderid', 'id', 'order number', 'order #'),
        'name': ('name', 'ship name', 'customer name', 'full name'),
        'address': ('address', 'street', 'street address', 'address 1', 'ship address'),
        'city': ('city', 'ship city'),
        'state': ('state', 'province', 'region', 'ship state'),
        'zip': ('zip', 'zipcode', 'postal code', 'ship zip', 'ship postal code'),
    },
}


def _present(value):
    return value is not None and str(value).strip() != ''


def _extract(normalized_order, schema):
    return {
        field: next((normalized_order.get(a) for a in aliases if _present(normalized_order.get(a))), None)
        for field, aliases in schema.items()
    }


def _build_views(orders):
    product_rows, shipping_rows = [], []
    for order in orders:
        normalized = {str(k).strip().lower(): v for k, v in (order or {}).items()}
        product_rows.append(_extract(normalized, FIELD_MAPS['product']))
        shipping_rows.append(_extract(normalized, FIELD_MAPS['shipping']))
    return product_rows, shipping_rows


def _save_active_orders_json(orders, batch_name='orders.json', source_file=None):
    JSON_DIR.mkdir(parents=True, exist_ok=True)
    out_path = JSON_DIR / 'orders.json'
    payload = {
        'generated_at': datetime.now().strftime('%Y%m%d_%H%M%S'),
        'batch_name': batch_name,
        'source_file': str(source_file) if source_file else None,
        'orders': orders,
    }
    with out_path.open('w', encoding='utf-8') as f:
        json.dump(payload, f, indent=2, ensure_ascii=True, default=str)
    return out_path


def processExcel(excel_path):
    excel_path = Path(excel_path)
    if not excel_path.exists():
        return None

    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    sheet = workbook.active

    row_iter = sheet.iter_rows(values_only=True)
    header_row = next(row_iter, None)
    if not header_row:
        return None

    headers = [str(cell).strip() if cell is not None else '' for cell in header_row]
    orders = []

    for row in row_iter:
        if row is None or all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        order = {
            (headers[i] if i < len(headers) and headers[i] else 'Column {}'.format(i + 1)): value
            for i, value in enumerate(row)
        }
        orders.append(order)

    product_rows, shipping_rows = _build_views(orders)

    batch_name = '{}.json'.format(excel_path.stem)
    active_path = _save_active_orders_json(orders, batch_name=batch_name, source_file=excel_path)

    print('Saved active orders: {}'.format(active_path))
    return {
        'json_file': str(active_path),
        'json_dir': str(JSON_DIR),
        'product': product_rows,
        'shipping': shipping_rows,
        'batch_name': batch_name,
    }


def _capture_with_takepicture(save_dir: Path) -> str | None:
    save_dir.mkdir(parents=True, exist_ok=True)
    print('Starting camera preview (takepicture). Press Enter to capture, Esc/q to cancel.')
    return takepicture.take_picture(save_dir=str(save_dir))


def _capture_with_autofocus(save_dir: Path, i2c_bus: int = 2, debug: bool = False) -> str | None:
    save_dir.mkdir(parents=True, exist_ok=True)
    captured = {'path': None}

    def _on_capture(image_path: str) -> bool:
        captured['path'] = image_path
        return True

    Autofocus.focuser = Focuser(i2c_bus, debug=debug)
    print('Starting camera preview (autofocus). Press Enter to capture, Esc to cancel.')
    show_camera(on_capture=_on_capture)

    if not captured['path']:
        return None

    output_path = save_dir / 'photo_{}.png'.format(datetime.now().strftime('%Y%m%d_%H%M%S'))
    shutil.move(captured['path'], output_path)
    print('Picture saved to {}'.format(output_path))
    return str(output_path)


def _select_order(orders: dict, order_index: int) -> dict | None:
    products = orders.get('product', []) if isinstance(orders, dict) else []
    if not products:
        return None
    if order_index < 0 or order_index >= len(products):
        return None
    return products[order_index]


def _normalize_expected_order(order_row: dict | None) -> dict:
    row = order_row if isinstance(order_row, dict) else {}
    return {
        'Name': '' if row.get('Name') is None else str(row.get('Name')),
        'Title': '' if row.get('Title') is None else str(row.get('Title')),
        'Date': '' if row.get('Date') is None else str(row.get('Date')),
    } if any(k in row for k in ('Name', 'Title', 'Date')) else {
        'Name': '' if row.get('name') is None else str(row.get('name')),
        'Title': '' if row.get('title') is None else str(row.get('title')),
        'Date': '' if row.get('date') is None else str(row.get('date')),
    }


def _evaluate_pass_fail(expected: dict, response_json: dict) -> dict:
    checks = {}
    overall_pass = True

    for field in ('Name', 'Title', 'Date'):
        expected_value = '' if expected.get(field) is None else str(expected.get(field))
        actual_value = '' if response_json.get(field) is None else str(response_json.get(field))
        matched = expected_value == actual_value
        checks[field] = {
            'pass': matched,
            'expected': expected_value,
            'actual': actual_value,
        }
        if not matched:
            overall_pass = False

    return {
        'overall_pass': overall_pass,
        'checks': checks,
    }


def _order_label(order_row: dict | None) -> str:
    row = order_row if isinstance(order_row, dict) else {}
    order_id = row.get('order_id', 'n/a')
    name = row.get('name') or row.get('Name') or ''
    return 'order_id={} name={}'.format(order_id, name)


def _print_order_result(evaluation: dict) -> None:
    if evaluation.get('overall_pass'):
        print('Result: PASS')
        return

    print('Result: FAIL')
    for field_name, status in evaluation.get('checks', {}).items():
        if status.get('pass'):
            continue
        print(
            "  {}: expected='{}' got='{}'".format(
                field_name,
                status.get('expected', ''),
                status.get('actual', ''),
            )
        )


def _build_fail_reasons(evaluation: dict) -> list[str]:
    reasons = []
    for field_name, status in evaluation.get('checks', {}).items():
        if status.get('pass'):
            continue
        reasons.append(
            "{} expected='{}' got='{}'".format(
                field_name,
                status.get('expected', ''),
                status.get('actual', ''),
            )
        )
    return reasons


def run(
    excel_path: str,
    use_autofocus: bool = False,
    order_index: int = 0,
    image_save_dir: str = 'photos',
    result_save_dir: str = 'results',
    api_key: str | None = None,
    document_template: list[str] | None = None,
) -> dict | None:
    orders = processExcel(excel_path)
    if not orders:
        print('Failed to load orders from: {}'.format(excel_path))
        return None

    selected_order = _select_order(orders, order_index)
    if not selected_order:
        print('No product row available for order_index={}'.format(order_index))
        return None

    print('Using {}'.format(_order_label(selected_order)))

    if use_autofocus:
        image_path = _capture_with_autofocus(Path(image_save_dir))
    else:
        image_path = _capture_with_takepicture(Path(image_save_dir))

    if not image_path:
        print('Image capture cancelled.')
        return None

    order_info = {'product': [selected_order]}
    analysis = groq.analyze_image_against_orders(
        order_info=order_info,
        image_path=image_path,
        api_key=api_key,
        save_json=False,
        document_template=document_template or DEFAULT_DOCUMENT_TEMPLATE,
    )

    expected = _normalize_expected_order(selected_order)
    response_json = analysis.get('response_json', {}) if isinstance(analysis, dict) else {}
    evaluation = _evaluate_pass_fail(expected, response_json)

    _print_order_result(evaluation)


    return {
        'orders': orders,
        'selected_order': selected_order,
        'image_path': image_path,
        'analysis': analysis,
        'evaluation': evaluation,
    }


def run_all(
    excel_path: str,
    use_autofocus: bool = False,
    image_save_dir: str = 'photos',
    result_save_dir: str = 'results',
    api_key: str | None = None,
    document_template: list[str] | None = None,
) -> dict | None:
    orders = processExcel(excel_path)
    if not orders:
        print('Failed to load orders from: {}'.format(excel_path))
        return None

    products = orders.get('product', []) if isinstance(orders, dict) else []
    if not products:
        print('No product rows found in {}'.format(excel_path))
        return None

    print('Processing {} order(s).'.format(len(products)))

    batch_results = []
    failed_orders = []
    pass_count = 0
    fail_count = 0

    for index, selected_order in enumerate(products):
        print('\n[{}/{}] {}'.format(index + 1, len(products), _order_label(selected_order)))

        if use_autofocus:
            image_path = _capture_with_autofocus(Path(image_save_dir))
        else:
            image_path = _capture_with_takepicture(Path(image_save_dir))

        if not image_path:
            print('Image capture cancelled. Stopping batch processing.')
            break

        order_info = {'product': [selected_order]}
        analysis = groq.analyze_image_against_orders(
            order_info=order_info,
            image_path=image_path,
            api_key=api_key,
            save_json=False,
            document_template=document_template or DEFAULT_DOCUMENT_TEMPLATE,
        )

        expected = _normalize_expected_order(selected_order)
        response_json = analysis.get('response_json', {}) if isinstance(analysis, dict) else {}
        evaluation = _evaluate_pass_fail(expected, response_json)

        _print_order_result(evaluation)

        if evaluation['overall_pass']:
            pass_count += 1
        else:
            fail_count += 1
            failed_orders.append(
                {
                    'order_id': selected_order.get('order_id', 'n/a') if isinstance(selected_order, dict) else 'n/a',
                    'reasons': _build_fail_reasons(evaluation),
                }
            )

        batch_results.append(
            {
                'order_index': index,
                'selected_order': selected_order,
                'image_path': image_path,
                'analysis': analysis,
                'evaluation': evaluation,
            }
        )

    summary = {
        'orders': orders,
        'processed_count': len(batch_results),
        'total_orders': len(products),
        'pass_count': pass_count,
        'fail_count': fail_count,
        'failed_orders': failed_orders,
        'batch_results': batch_results,
    }

    print('\nSummary: processed {}/{} | pass={} fail={}'.format(
        summary['processed_count'],
        summary['total_orders'],
        summary['pass_count'],
        summary['fail_count'],
    ))
    if failed_orders:
        print('Failed orders:')
        for item in failed_orders:
            print('  order_id={}'.format(item['order_id']))
            for reason in item['reasons']:
                print('    - {}'.format(reason))

    return summary