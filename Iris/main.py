import argparse
import json
import logging
import shlex
from datetime import datetime
from functools import partial
from pathlib import Path

import Autofocus
import Image_llm
from Autofocus import show_camera
from Focuser import Focuser
from openpyxl import load_workbook

EXCEL_PATH = Path(r"C:\Users\EvanN\Downloads\PrintProof\OCR\PrintProof\orders.xlsx")
JSON_DIR = Path(r"C:\Users\EvanN\Downloads\PrintProof\OCR\PrintProof\Iris\Json")
BATCH_DIR = Path(r"C:\Users\EvanN\Downloads\PrintProof\OCR\PrintProof\Iris\batch")
LOG_DIR = Path(r"C:\Users\EvanN\Downloads\PrintProof\OCR\PrintProof\Iris\logs")
LOG_FILE = LOG_DIR / 'iris.log'
FIELD_MAPS = {
    'product': {
        'order_id': ('order id', 'orderid', 'id', 'order number', 'order #'),
        'name': ('name', 'customer name', 'ship name', 'full name'),
        'title': ('title', 'product title', 'item title', 'item', 'product'),
        'date': ('date', 'order date', 'purchase date', 'created at'),
    },
    'shipping': {
        'order_id': ('order id', 'orderid', 'id', 'order number', 'order #'),
        'name': ('name', 'ship name', 'customer name', 'full name'),
        'address': ('address', 'street', 'street address', 'address 1', 'ship address'),
        'city': ('city', 'ship city'),
        'state': ('state', 'province', 'region', 'ship state'),
        'zip': ('zip', 'zipcode', 'postal code', 'ship zip', 'ship postal code'),
    },
}


def _present(value):
    return value is not None and str(value).strip() != ''


def _extract(normalized_order, schema):
    return {
        field: next((normalized_order.get(a) for a in aliases if _present(normalized_order.get(a))), None)
        for field, aliases in schema.items()
    }


def _build_views(orders):
    product_rows, shipping_rows = [], []
    for order in orders:
        normalized = {str(k).strip().lower(): v for k, v in (order or {}).items()}
        product_rows.append(_extract(normalized, FIELD_MAPS['product']))
        shipping_rows.append(_extract(normalized, FIELD_MAPS['shipping']))
    return product_rows, shipping_rows


def _save_active_orders_json(orders, batch_name='orders.json', source_file=None):
    JSON_DIR.mkdir(parents=True, exist_ok=True)
    out_path = JSON_DIR / 'orders.json'
    payload = {
        'generated_at': datetime.now().strftime('%Y%m%d_%H%M%S'),
        'batch_name': batch_name,
        'source_file': str(source_file) if source_file else None,
        'orders': orders,
    }
    with out_path.open('w', encoding='utf-8') as f:
        json.dump(payload, f, indent=2, ensure_ascii=True, default=str)
    return out_path


def processExcel(excel_path):
    excel_path = Path(excel_path)
    if not excel_path.exists():
        return None

    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    sheet = workbook.active

    row_iter = sheet.iter_rows(values_only=True)
    header_row = next(row_iter, None)
    if not header_row:
        return None

    headers = [str(cell).strip() if cell is not None else '' for cell in header_row]
    orders = []

    for row in row_iter:
        if row is None or all(cell is None or str(cell).strip() == '' for cell in row):
            continue

        order = {
            (headers[i] if i < len(headers) and headers[i] else 'Column {}'.format(i + 1)): value
            for i, value in enumerate(row)
        }
        orders.append(order)

    product_rows, shipping_rows = _build_views(orders)

    batch_name = '{}.json'.format(excel_path.stem)
    active_path = _save_active_orders_json(orders, batch_name=batch_name, source_file=excel_path)

    print('Saved active orders: {}'.format(active_path))
    return {
        'json_file': str(active_path),
        'json_dir': str(JSON_DIR),
        'product': product_rows,
        'shipping': shipping_rows,
        'batch_name': batch_name,
    }


def loadBatch(batch_name):
    BATCH_DIR.mkdir(parents=True, exist_ok=True)
    batch_path = BATCH_DIR / batch_name
    if not batch_path.exists():
        return None

    with batch_path.open('r', encoding='utf-8') as f:
        payload = json.load(f)

    orders = payload.get('orders', []) if isinstance(payload, dict) else []
    _save_active_orders_json(orders, batch_name=batch_path.name, source_file=payload.get('source_file') if isinstance(payload, dict) else None)
    product_rows, shipping_rows = _build_views(orders)
    return {
        'json_file': str(batch_path),
        'json_dir': str(BATCH_DIR),
        'product': product_rows,
        'shipping': shipping_rows,
        'batch_name': batch_path.name,
    }


def setup_logging():
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        filename=str(LOG_FILE),
        level=logging.INFO,
        format='%(asctime)s %(levelname)s %(message)s',
    )


def _filter_orders(orders, order_id=None):
    products = (orders or {}).get('product', []) if isinstance(orders, dict) else []
    if order_id is None:
        return {'product': products}
    needle = str(order_id).strip().lower()
    return {
        'product': [
            row for row in products
            if str((row or {}).get('order_id', '')).strip().lower() == needle
        ]
    }

def parse_cmdline():
    parser = argparse.ArgumentParser(description='PrintProof — LLM-only mode (no OCR)')
    parser.add_argument('-i', '--i2c-bus', type=int, default=2,
                        help='Set i2c bus (A02=6, B01=7/8, Xavier NX=9/10)')
    parser.add_argument('--debug', action='store_true',
                        help='Print focus write operations for debugging')
    return parser.parse_args()


def handle_capture(image_path, orders=None):
    result = Image_llm.ProcessImage(image_path, orders)
    logging.info('capture image=%s products=%d result_len=%d', image_path, len((orders or {}).get('product', [])), len(result or ''))
    return result


def _load_orders_source(source):
    source_path = Path(source)
    if source_path.suffix.lower() == '.xlsx':
        return processExcel(source_path)
    if source_path.suffix.lower() == '.json':
        if source_path.is_absolute():
            if not source_path.exists():
                return None
            with source_path.open('r', encoding='utf-8') as f:
                payload = json.load(f)
            orders = payload.get('orders', []) if isinstance(payload, dict) else []
            _save_active_orders_json(
                orders,
                batch_name=source_path.name,
                source_file=payload.get('source_file') if isinstance(payload, dict) else None,
            )
            product_rows, shipping_rows = _build_views(orders)
            return {
                'json_file': str(source_path),
                'json_dir': str(source_path.parent),
                'product': product_rows,
                'shipping': shipping_rows,
                'batch_name': source_path.name,
            }
        return loadBatch(source_path.name)
    return None


def run_terminal(args, orders):
    state = {
        'orders': orders or {'product': [], 'shipping': []},
        'last_action': None,
        'last_result': None,
        'batch_name': (orders or {}).get('batch_name') if isinstance(orders, dict) else None,
    }

    print('Command terminal ready. Type help for commands.')
    while True:
        try:
            raw = input('iris> ').strip()
        except (EOFError, KeyboardInterrupt):
            print('')
            break
        if not raw:
            continue

        parts = shlex.split(raw)
        cmd = parts[0].lower()
        logging.info('cmd: %s', raw)

        if cmd in ('exit', 'quit'):
            break

        if cmd == 'help':
            print(
                'Commands:\n'
                '  help\n'
                '  list\n'
                '  batch <full_or_relative_file_path>\n'
                '  run <image_path>\n'
                '  runid <order_id> <image_path>\n'
                '  camera\n'
                '  rerun\n'
                '  reload\n'
                '  quit'
            )
            continue

        if cmd == 'list':
            products = state['orders'].get('product', [])
            for row in products[:50]:
                print('order_id={} title={} name={}'.format(row.get('order_id'), row.get('title'), row.get('name')))
            print('total product rows: {}'.format(len(products)))
            continue

        if cmd == 'reload':
            if not state['batch_name']:
                print('No active batch. Use: batch <path-to-orders.xlsx-or-json>')
                continue

            loaded = loadBatch(state['batch_name'])
            if loaded:
                state['orders'] = loaded
                print('Reloaded batch from {}'.format(state['batch_name']))
            else:
                print('Failed to reload active batch: {}'.format(state['batch_name']))
            continue

        if cmd == 'batch' and len(parts) == 2:
            batch_input = parts[1]
            loaded = _load_orders_source(batch_input)
            if Path(batch_input).suffix.lower() not in ('.xlsx', '.json'):
                print('Batch command expects .xlsx or .json, e.g. batch ./path/to/orders.xlsx')
                continue

            if not loaded:
                print('Failed to load batch from: {}'.format(batch_input))
                continue
            state['orders'] = loaded
            state['batch_name'] = loaded.get('batch_name', state['batch_name'])
            print('Loaded batch: {}'.format(state['batch_name']))
            Autofocus.focuser = Focuser(args.i2c_bus, debug=args.debug)
            show_camera(on_capture=partial(handle_capture, orders=state['orders']))
            continue

        if cmd == 'run' and len(parts) == 2:
            image_path = parts[1]
            filtered = _filter_orders(state['orders'])
            state['last_action'] = ('run', image_path, None)
            state['last_result'] = Image_llm.ProcessImage(image_path, filtered)
            continue

        if cmd == 'runid' and len(parts) == 3:
            order_id, image_path = parts[1], parts[2]
            filtered = _filter_orders(state['orders'], order_id)
            if not filtered['product']:
                print('No matching order_id: {}'.format(order_id))
                continue
            state['last_action'] = ('runid', image_path, order_id)
            state['last_result'] = Image_llm.ProcessImage(image_path, filtered)
            continue

        if cmd == 'rerun':
            if not state['last_action']:
                print('Nothing to rerun yet.')
                continue
            action, image_path, order_id = state['last_action']
            filtered = _filter_orders(state['orders'], order_id)
            state['last_result'] = Image_llm.ProcessImage(image_path, filtered)
            continue

        if cmd == 'camera':
            Autofocus.focuser = Focuser(args.i2c_bus, debug=args.debug)
            show_camera(on_capture=partial(handle_capture, orders=state['orders']))
            continue

        print('Unknown command. Type help.')


if __name__ == '__main__':
    setup_logging()
    args = parse_cmdline()
    orders = None
    run_terminal(args, orders)